import datetime as dt

from core.fakes import fake_hq
from tracker import migrate

ASHBY = "https://jobs.ashbyhq.com/plaid/64642463-8ee8-4453-b166-449711f61cd0"


# ------------------------------------------------------------------ legacy

def _legacy_row(id_, **kw):
    base = {"id": id_, "company": "Acme", "title": "PM", "location": "NYC",
            "url": "https://boards.greenhouse.io/acme/jobs/1", "status": "New",
            "first_seen": "2026-06-01", "last_seen": "2026-07-01", "posted": "",
            "yoe": "5+", "seniority": "Senior", "company_industry": "Fintech",
            "role_focus": "Payments", "skills": "APIs; SQL", "comp_range": "$180k",
            "work_model": "Remote", "tagged_at": "2026-06-02"}
    base.update(kw)
    return base


def test_legacy_appends_missing_preserves_fields(monkeypatch):
    from monitor import tagging
    monkeypatch.setattr(tagging, "min_yoe_from", lambda y: 5 if y == "5+" else "",
                        raising=False)
    hq = fake_hq()
    hq.tab("feed").append_records([{"key": "greenhouse-1", "company": "Old"}])
    rows = [_legacy_row("greenhouse-1"),                     # already there
            _legacy_row("lever-aaaaaaaa-1111-2222-3333-444444444444", status="Closed"),
            _legacy_row("")]                                 # unkeyable
    rep = migrate.migrate_legacy(hq, rows)
    assert rep == {"source_rows": 3, "appended": 1, "skipped": 2}
    rec = [r for r in hq.tab("feed").records() if r["key"].startswith("lever-")][0]
    assert rec["status"] == "Closed"
    assert rec["first_seen"] == "2026-06-01" and rec["tagged_at"] == "2026-06-02"
    assert rec["skills"] == "APIs; SQL"
    assert rec["min_yoe"] == "5"                             # via min_yoe_from


def test_legacy_min_yoe_blank_when_helper_missing(monkeypatch):
    from monitor import tagging
    monkeypatch.delattr(tagging, "min_yoe_from", raising=False)
    hq = fake_hq()
    migrate.migrate_legacy(hq, [_legacy_row("greenhouse-9")])
    assert hq.tab("feed").records()[0]["min_yoe"] == ""


def test_legacy_dry_run_writes_nothing():
    hq = fake_hq()
    rep = migrate.migrate_legacy(hq, [_legacy_row("greenhouse-2")], dry_run=True)
    assert rep["appended"] == 1
    assert hq.tab("feed").records() == []


# -------------------------------------------------------------- scout xlsx

def _mini_workbook(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs Applied"
    ws.cell(row=1, column=2, value="Shan Shaheen")          # junk banner rows
    hdr = ["Company", "City", "Position", "Remote / Hybrid /Onsite",
           "Date Searched (mm/dd/yyyy)", "Min Experience Required",
           "Job Link", "email", "abc"]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=i, value=h)
    data = [
        ["Abridge", "SF", "Product Ops", "Remote", dt.datetime(2026, 7, 3), "2 Yrs",
         ASHBY, "", "x"],
        ["Oldco", "NY", "PM", "", dt.datetime(2025, 12, 1), "", "", "", ""],
        ["", "", "", "", dt.datetime(2026, 7, 3), "", "", "", ""],
        ["Abridge", "SF", "Product Ops", "Remote", dt.datetime(2026, 7, 4), "2 Yrs",
         ASHBY, "", ""],                                     # dup of row 1 by key
        ["StrDate", "Chicago", "PM II", "Hybrid", "07/05/2026", "3 Yrs", "", "", ""],
    ]
    for r, row in enumerate(data, start=5):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return str(path)


def test_scout_xlsx_maps_noisy_headers_filters_2026(tmp_path):
    hq = fake_hq()
    path = _mini_workbook(tmp_path / "scout.xlsx")
    rep = migrate.migrate_scout_xlsx(hq, path)
    assert rep == {"appended": 2, "skipped_dup": 1, "skipped_pre2026": 1}
    recs = {r["Company"]: r for r in hq.tab("scout_jobs").records()}
    ab = recs["Abridge"]
    assert ab["Applied"] == "TRUE"
    assert ab["Remote / Hybrid / Onsite"] == "Remote"       # noisy header mapped
    assert ab["Min Experience"] == "2 Yrs"
    assert ab["Date Searched"] == "2026-07-03"
    assert ab["Job Link"] == ASHBY
    assert recs["StrDate"]["Date Searched"] == "2026-07-05"  # string date parsed


def test_scout_xlsx_idempotent_and_missing_file(tmp_path):
    hq = fake_hq()
    path = _mini_workbook(tmp_path / "scout.xlsx")
    migrate.migrate_scout_xlsx(hq, path)
    rep2 = migrate.migrate_scout_xlsx(hq, path)
    assert rep2["appended"] == 0
    assert len(hq.tab("scout_jobs").records()) == 2
    assert migrate.migrate_scout_xlsx(hq, str(tmp_path / "nope.xlsx"))["appended"] == 0


# ----------------------------------------------------------------- applog

def _applog(tmp_path) -> str:
    p = tmp_path / "applications-log.csv"
    p.write_text(
        "timestamp,slug,event,company,role,url,file\n"
        f"2026-07-06T23:54,plaid,created,Plaid,Product Manager,{ASHBY},\n"
        "2026-07-06T23:56,plaid,packaged,Plaid,,,resume.pdf\n"
        "2026-07-07T01:00,acme,created,Acme,PM,https://boards.greenhouse.io/acme/jobs/777,\n"
        "2026-07-08T09:00,acme,applied,Acme,,,\n"
        "2026-07-09T00:00,zeta,created,Zeta,PM Platform,,\n")
    return str(p)


def test_applog_groups_by_slug_and_keys_from_created_row(tmp_path):
    hq = fake_hq()
    rep = migrate.migrate_applog(hq, _applog(tmp_path))
    assert rep["slugs"] == 3
    assert rep["appended"] == 2          # zeta never applied/packaged
    recs = {r["company"]: r for r in hq.tab("pipeline").records()}
    assert recs["Acme"]["key"] == "greenhouse-777"
    assert recs["Acme"]["status"] == "Applied"
    assert recs["Acme"]["source"] == "manual"
    assert recs["Acme"]["applied_via"] == "self"
    assert recs["Acme"]["applied_date"] == "2026-07-08"
    assert recs["Plaid"]["title"] == "Product Manager"       # merged from created row
    assert recs["Plaid"]["applied_date"] == "2026-07-06"


def test_applog_skips_existing_keys_and_missing_file(tmp_path):
    hq = fake_hq()
    hq.tab("pipeline").append_records([{"key": "greenhouse-777", "company": "Acme"}])
    rep = migrate.migrate_applog(hq, _applog(tmp_path))
    assert rep["appended"] == 1 and rep["skipped_dup"] == 1
    rep2 = migrate.migrate_applog(hq, str(tmp_path / "gone" / "applications-log.csv"))
    assert rep2["appended"] == 0 and "missing" in rep2


def test_applog_dry_run(tmp_path):
    hq = fake_hq()
    rep = migrate.migrate_applog(hq, _applog(tmp_path), dry_run=True)
    assert rep["appended"] == 2
    assert hq.tab("pipeline").records() == []


_SIMPLIFY_CSV = (
    "Job Title,Company Name,Location,Job URL,Applied Date,Status,Status Date,"
    "Archived,Date Archived,Notes,job_type\n"
    "PM,Garner Health,NY,https://job-boards.greenhouse.io/garnerhealth/jobs/6101768004,"
    "2026-07-13,APPLIED,2026-07-13,No,N/A,,Full-time\n"
    "PM,Plaid,SF,https://jobs.ashbyhq.com/plaid/8d2c62be-4fce-4e69-8335-6f15a2553b62,"
    "N/A,SAVED,2026-07-10,No,N/A,,Full-time\n"
    "PM,Ramp,NY,https://jobs.ashbyhq.com/ramp/aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee,"
    "2026-07-01,INTERVIEWING,2026-07-09,No,N/A,note here,Full-time\n"
)


def _simplify(tmp_path):
    p = tmp_path / "simplify.csv"
    p.write_text(_SIMPLIFY_CSV)
    return str(p)


def test_simplify_csv_maps_status_source_and_applied(tmp_path):
    hq = fake_hq()
    rep = migrate.migrate_simplify_csv(hq, _simplify(tmp_path), dry_run=False)
    assert rep == {"rows": 3, "created": 3, "advanced_existing": 0, "unkeyed": 0}
    rows = {r["company"]: r for r in hq.tab("pipeline").records()}
    assert rows["Garner Health"]["status"] == "Applied"
    assert rows["Garner Health"]["source"] == "simplify"
    assert rows["Garner Health"]["applied_via"] == "simplify"
    assert rows["Garner Health"]["applied_date"] == "2026-07-13"
    assert rows["Plaid"]["status"] == "Inbox"          # SAVED -> Inbox, not applied
    assert rows["Plaid"]["applied_via"] == ""
    assert rows["Ramp"]["status"] == "Interview"
    assert rows["Ramp"]["notes"] == "note here"


def test_simplify_csv_idempotent_and_forward_only(tmp_path):
    hq = fake_hq()
    path = _simplify(tmp_path)
    migrate.migrate_simplify_csv(hq, path, dry_run=False)
    # a human already moved Garner to Offer — the re-import must not downgrade it
    hq.tab("pipeline").set_by_key(
        [r["key"] for r in hq.tab("pipeline").records() if r["company"] == "Garner Health"][0],
        {"status": "Offer"})
    rep = migrate.migrate_simplify_csv(hq, path, dry_run=False)
    assert rep["created"] == 0 and rep["advanced_existing"] == 3
    assert len(hq.tab("pipeline").records()) == 3        # no duplicates
    garner = [r for r in hq.tab("pipeline").records() if r["company"] == "Garner Health"][0]
    assert garner["status"] == "Offer"                   # forward-only, human wins


def test_simplify_csv_missing_file_is_noop(tmp_path):
    hq = fake_hq()
    rep = migrate.migrate_simplify_csv(hq, str(tmp_path / "nope.csv"), dry_run=False)
    assert rep["created"] == 0 and "missing" in rep
